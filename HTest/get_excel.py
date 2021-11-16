#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# 读取excel文件
import openpyxl


class Getexcel:
    def __init__(self, filepath):
        self.path = filepath
        
    def read_excel(self):
        """
        读取Excel文件
        :return: 包含多个列表的列表，例如 [[],[],[]……]
        order by taolei
        """
        table = openpyxl.load_workbook(self.path)
        sheet_name = table.sheetnames
        sheet = table[sheet_name[0]]  # 默认读取第一个sheet页
        rows = sheet.rows
        sheet_list = []
        for row in rows:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            sheet_list.append(row_list)
        return sheet_list
    
    
    def write_excel(self, data: list):
        """
        按行写入excel文件
        :param data: [["name", "age"], ["lily", 18], ["lucy", 20]]
        """
        excel = openpyxl.Workbook()
        sheet = excel.create_sheet('sheet1', 0)  # 创建sheet，并指定为第一个sheet
        for row_no, row in enumerate(data):
            for col_no, value in enumerate(row):
                sheet.cell(row_no + 1, col_no + 1).value = value
        excel.save(self.path)
